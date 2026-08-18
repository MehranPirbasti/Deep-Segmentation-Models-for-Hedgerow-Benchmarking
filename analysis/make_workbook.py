"""Puts all the figure data in one spreadsheet with a legend."""
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import tables as T

M = pd.read_csv("../data/master.csv")
wb = Workbook(); wb.remove(wb.active)
HEAD = Font(name="Arial", size=10, bold=True, color="FFFFFF")
BODY = Font(name="Arial", size=10)
BLUE = Font(name="Arial", size=10, color="0000FF")
FILL = PatternFill("solid", fgColor="1F3864")
LOCK = PatternFill("solid", fgColor="EDEDED")
THIN = Border(bottom=Side(style="thin", color="BFBFBF"))


def sheet(name, df, note, editable=(), locked=()):
    ws = wb.create_sheet(name[:31])
    ws["A1"] = note
    ws["A1"].font = Font(name="Arial", size=10, italic=True, color="555555")
    ws.append([])
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    hdr = 3
    for c in range(1, df.shape[1] + 1):
        cell = ws.cell(hdr, c); cell.font = HEAD; cell.fill = FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = max(
            12, min(26, int(df.columns[c-1].__len__()) + 6))
    for r in range(hdr + 1, hdr + 1 + len(df)):
        for c in range(1, df.shape[1] + 1):
            col = df.columns[c-1]
            cell = ws.cell(r, c); cell.border = THIN
            if col in editable:
                cell.font = BLUE
            else:
                cell.font = BODY
                if col in locked:
                    cell.fill = LOCK
            if isinstance(cell.value, float):
                cell.number_format = "0.000" if col == "loss" else "0.00"
    ws.freeze_panes = ws.cell(hdr + 1, 1)
    return ws


# legend
ws = wb.create_sheet("READ ME")
rows = [
 ("Figure data for TGRS-2026-02906", ""),
 ("", ""),
 ("Every figure in the manuscript is drawn from these tables by build_v6.py.", ""),
 ("Edit a value here, re-run the script, and the figure changes accordingly.", ""),
 ("", ""),
 ("Blue text", "values you may edit"),
 ("Grey fill", "values fixed by a manuscript table; editing one will make the "
               "figure disagree with the paper, and audit.py will report it"),
 ("", ""),
 ("Rebuild everything", "python master.py && python build_v6.py -out Fig -topology"),
 ("Verify against the tables", "python audit.py"),
 ("Statistical checks", "python stats_audit.py"),
 ("", ""),
 ("Sheet", "Contents"),
 ("ALL_CONFIGURATIONS", "the single source table; every other sheet is a view of it"),
 ("Table_II_III / V / VI / VII", "the manuscript tables, which take precedence"),
 ("DecoderBoxplot, LearningCurves", "series recovered from the original submission figures"),
 ("PROVENANCE", "where each column comes from"),
]
for i, (a, b) in enumerate(rows, start=1):
    ws.cell(i, 1, a).font = Font(name="Arial", size=11,
                                 bold=(i == 1 or a in ("Sheet",)), size2=None) if False else Font(
        name="Arial", size=11, bold=(i == 1 or a == "Sheet"))
    ws.cell(i, 2, b).font = BODY
ws.column_dimensions["A"].width = 30; ws.column_dimensions["B"].width = 96
ws["A6"].font = BLUE; ws["A7"].fill = LOCK

TABLE_COLS = {"iou","dice_f1","boundary_f1","inference_time_s","train_hours"}
ALL = ["Model","Decoder","Backbone","Group","iou","dice_f1","boundary_f1","bf1_r1",
       "cldice","betti0_err","frag_index","inference_time_s","train_hours","loss",
       "topo_measured","src"]
sheet("ALL_CONFIGURATIONS", M[ALL],
      "One row per configuration. 'src' says whether the row is fixed by a "
      "manuscript table or recovered from an original figure.",
      editable=[c for c in ALL if c not in TABLE_COLS],
      locked=list(TABLE_COLS))
sheet("Table_II_III", T.table_df(),
      "Tables II and III of the manuscript, verbatim. Authoritative.",
      locked=list(T.table_df().columns))
sheet("Table_V_seeds", T.seeds(),
      "Table V. Mean, 95 % CI half-width and SD of test IoU over five seeds.",
      locked=["Model","iou_mean","ci95","sd"])
topo = pd.DataFrame([(f"{d} ({b})", *v) for (d, b), v in T.TOPOLOGY.items()],
                    columns=["Model","cldice","betti0_err","frag_index"])
sheet("Table_VI_topology", topo, "Table VI. The nine measured configurations.",
      locked=list(topo.columns))
bf = pd.DataFrame([(f"{d} ({b})", v) for (d, b), v in T.BF1_R1.items()],
                  columns=["Model","bf1_r1"])
bf["boundary_f1"] = [float(M[M.Model == m].boundary_f1.iloc[0]) for m in bf.Model]
bf["delta"] = (bf.boundary_f1 - bf.bf1_r1).round(2)
sheet("Table_VII_bf1_tol", bf, "Table VII. BF1 at the two tolerance radii.",
      locked=["Model","bf1_r1","boundary_f1"])
sheet("DecoderBoxplot", pd.read_csv("../data/decoder_val_iou.csv"),
      "Best validation IoU per configuration, recovered from the "
      "decoder-distribution figure of the original submission.")
lc = pd.read_csv("../data/convergence_bands.csv")
sheet("LearningCurves", lc.round(4),
      "Per-encoder-family min, median and max validation IoU by epoch, "
      "recovered from the convergence figure of the original submission.")
prov = pd.read_csv("../data/_provenance.csv")
sheet("PROVENANCE", prov, "Where each quantity comes from.")

wb.save("../data/Figure_Data.xlsx")
print("wrote Figure_Data.xlsx with", len(wb.sheetnames), "sheets:", wb.sheetnames)
